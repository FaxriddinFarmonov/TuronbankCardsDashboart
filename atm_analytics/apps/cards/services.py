# from openpyxl import load_workbook
# from .models import ActiveCard
#
# from datetime import datetime
# from decimal import Decimal, InvalidOperation
#
# import logging
#
#
# logger = logging.getLogger(__name__)
#
# # =========================================
# # SETTINGS
# # =========================================
# BATCH_SIZE = 2000
#
#
# # =========================================
# # DATE PARSER
# # =========================================
# def parse_date(value):
#     """
#     Excel ichidagi sanani xavfsiz parse qiladi.
#     """
#
#     if value is None or value == "":
#         return None
#
#     # Agar datetime object bo‘lsa
#     if isinstance(value, datetime):
#         return value.date()
#
#     value = str(value).strip()
#
#     if not value:
#         return None
#
#     # mumkin bo'lgan formatlar
#     formats = [
#         "%d.%m.%Y",
#         "%Y-%m-%d",
#         "%d/%m/%Y",
#     ]
#
#     for fmt in formats:
#         try:
#             return datetime.strptime(value, fmt).date()
#         except Exception:
#             continue
#
#     return None
#
#
# # =========================================
# # DECIMAL PARSER
# # =========================================
# def parse_decimal(value):
#     """
#     Decimal qiymatlarni xavfsiz parse qiladi.
#     """
#
#     if value is None or value == "":
#         return Decimal("0")
#
#     # int yoki float bo‘lsa
#     if isinstance(value, (int, float)):
#         return Decimal(str(value))
#
#     value = str(value).strip()
#
#     if not value:
#         return Decimal("0")
#
#     # probellarni olib tashlash
#     value = value.replace(" ", "")
#
#     # vergulni nuqtaga aylantirish
#     value = value.replace(",", ".")
#
#     try:
#         return Decimal(value)
#     except (InvalidOperation, ValueError):
#         return Decimal("0")
#
#
# # =========================================
# # STRING CLEANER
# # =========================================
# def clean_string(value):
#     """
#     String qiymatlarni tozalaydi.
#     """
#
#     if value is None:
#         return ""
#
#     return str(value).strip()
#
#
# # =========================================
# # MAIN IMPORT FUNCTION
# # =========================================
# def import_excel(file_path):
#     """
#     Katta Excel fayllarni tez va xavfsiz import qiladi.
#     """
#
#     wb = load_workbook(
#         filename=file_path,
#         read_only=True,
#         data_only=True
#     )
#
#     ws = wb.active
#
#     objects = []
#
#     success_count = 0
#     error_count = 0
#
#     # ==================================================
#     # HEADER SKIP
#     # ==================================================
#     rows = ws.iter_rows(
#         min_row=4,
#         values_only=True
#     )
#
#     for index, row in enumerate(rows, start=4):
#
#         try:
#             # bo'sh row skip
#             if not row:
#                 continue
#
#             # row length tekshirish
#             if len(row) < 19:
#                 error_count += 1
#                 continue
#
#             # account bo‘sh bo‘lsa skip
#             if not row[5]:
#                 continue
#
#             obj = ActiveCard(
#
#                 # 1
#                 row_number=int(row[0]) if row[0] else 0,
#
#                 # 2
#                 branch_code=clean_string(row[1]),
#
#                 # 3
#                 branch_name=clean_string(row[2]),
#
#                 # 4
#                 parent_branch=clean_string(row[3]),
#
#                 # 5
#                 level=clean_string(row[4]),
#
#                 # 6
#                 account=clean_string(row[5]),
#
#                 # 7
#                 dt_turnover=parse_date(row[6]),
#                 ct_turnover=parse_date(row[7]),
#
#                 # 9
#                 client_type=clean_string(row[8]),
#
#                 # 10
#                 card_type=clean_string(row[9]),
#
#                 # 11
#                 balance=parse_decimal(row[10]),
#
#                 # 12
#                 balance_equivalent=parse_decimal(row[11]),
#
#                 # 13
#                 doc_date=parse_date(row[12]),
#
#                 # 14
#                 expire_date=parse_date(row[13]),
#
#                 # 15
#                 card_system=clean_string(row[14]),
#
#                 # 16
#                 card_status=clean_string(row[15]),
#
#                 # 17
#                 currency_code=clean_string(row[16]),
#
#                 # 18
#                 resident_status=clean_string(row[17]),
#
#                 # 19
#                 client_full_name=clean_string(row[18]),
#             )
#
#             objects.append(obj)
#
#             success_count += 1
#
#             # =========================================
#             # BULK INSERT
#             # =========================================
#             if len(objects) >= BATCH_SIZE:
#
#                 ActiveCard.objects.bulk_create(
#                     objects,
#                     batch_size=BATCH_SIZE
#                 )
#
#                 objects.clear()
#
#         except Exception as e:
#
#             error_count += 1
#
#             logger.error(
#                 f"Excel import error | Row: {index} | Error: {str(e)}"
#             )
#
#             continue
#
#     # =========================================
#     # FINAL SAVE
#     # =========================================
#     if objects:
#         ActiveCard.objects.bulk_create(
#             objects,
#             batch_size=BATCH_SIZE
#         )
#
#     wb.close()
#
#     logger.info(
#         f"""
#         Excel import finished.
#
#         SUCCESS: {success_count}
#         ERRORS: {error_count}
#         """
#     )
#
#     return {
#         "success": success_count,
#         "errors": error_count,
#     }

from openpyxl import load_workbook
from .models import ActiveCard
from .branch_mapping import BRANCH_MAPPING

from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

import logging


logger = logging.getLogger(__name__)


# =========================================================
# SETTINGS
# =========================================================
BATCH_SIZE = 10000


# =========================================================
# STRING CLEANER
# =========================================================
def clean_string(value):
    """
    String qiymatlarni xavfsiz tozalash
    """

    if value is None:
        return ""

    return str(value).strip()


# =========================================================
# DATE PARSER
# =========================================================
def parse_date(value):
    """
    Excel date parser
    """
    if value is None or value == "":
        return None

    # datetime object
    if isinstance(value, datetime):
        return value.date()

    value = str(value).strip()

    if not value:
        return None

    formats = [
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]

    for fmt in formats:

        try:
            return datetime.strptime(
                value,
                fmt
            ).date()

        except Exception:
            continue

    return None


# =========================================================
# DECIMAL PARSER
# =========================================================
def parse_decimal(value):
    """
    Decimal parser
    """

    if value is None or value == "":
        return Decimal("0")

    # int/float
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    value = str(value).strip()

    if not value:
        return Decimal("0")

    # remove spaces
    value = value.replace(" ", "")

    # comma -> dot
    value = value.replace(",", ".")

    try:
        return Decimal(value)

    except (InvalidOperation, ValueError):
        return Decimal("0")


# =========================================================
# BRANCH INFO
# =========================================================
def get_branch_info(branch_code):
    """
    Branch mapping
    """

    branch_code = clean_string(branch_code)

    branch_info = BRANCH_MAPPING.get(branch_code)

    if not branch_info:

        return {
            "id_bank": "",
            "head_office": "",
        }

    return {
        "id_bank": clean_string(
            branch_info.get("id_bank")
        ),

        "head_office": clean_string(
            branch_info.get("head_office")
        ),
    }


# =========================================================
# MAIN IMPORT
# =========================================================
def import_excel(file_path):
    """
    Ultra Fast Excel Import
    """


    wb = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True
    )

    ws = wb.active

    rows = ws.iter_rows(
        min_row=4,
        values_only=True
    )

    objects = []

    success_count = 0
    error_count = 0

    # =====================================================
    # TRANSACTION
    # =====================================================
    with transaction.atomic():

        for index, row in enumerate(rows, start=4):

            try:

                # =============================================
                # EMPTY ROW
                # =============================================
                if not row:
                    continue

                # =============================================
                # COLUMN CHECK
                # =============================================
                if len(row) < 19:
                    error_count += 1
                    continue

                # =============================================
                # ACCOUNT CHECK
                # =============================================
                if not row[5]:
                    continue

                # =============================================
                # BRANCH
                # =============================================
                branch_code = clean_string(row[1])

                branch_info = get_branch_info(
                    branch_code
                )

                # =============================================
                # OBJECT
                # =============================================
                obj = ActiveCard(

                    # 1
                    row_number=int(row[0]) if row[0] else 0,

                    # 2
                    branch_code=branch_code,

                    # NEW
                    id_bank=branch_info["id_bank"],

                    # NEW
                    head_office=branch_info["head_office"],

                    # 3
                    branch_name=clean_string(row[2]),

                    # 4
                    parent_branch=clean_string(row[3]),

                    # 5
                    level=clean_string(row[4]),

                    # 6
                    account=clean_string(row[5]),

                    # 7
                    dt_turnover=parse_date(row[6]),

                    # 8
                    ct_turnover=parse_date(row[7]),

                    # 9
                    client_type=clean_string(row[8]),

                    # 10
                    card_type=clean_string(row[9]),

                    # 11
                    balance=parse_decimal(row[10]),

                    # 12
                    balance_equivalent=parse_decimal(
                        row[11]
                    ),

                    # 13
                    doc_date=parse_date(row[12]),

                    # 14
                    expire_date=parse_date(row[13]),

                    # 15
                    card_system=clean_string(row[14]),

                    # 16
                    card_status=clean_string(row[15]),

                    # 17
                    currency_code=clean_string(row[16]),

                    # 18
                    resident_status=clean_string(row[17]),

                    # 19
                    client_full_name=clean_string(
                        row[18]
                    ),
                )

                objects.append(obj)

                success_count += 1

                # =============================================
                # BULK INSERT
                # =============================================
                if len(objects) >= BATCH_SIZE:

                    ActiveCard.objects.bulk_create(
                        objects,
                        batch_size=BATCH_SIZE
                    )

                    # MEMORY CLEAR
                    objects.clear()

            except Exception as e:

                error_count += 1

                logger.error(
                    f"ROW {index} ERROR: {str(e)}"
                )

                continue

        # =================================================
        # FINAL INSERT
        # =================================================
        if objects:

            ActiveCard.objects.bulk_create(
                objects,
                batch_size=BATCH_SIZE
            )

    wb.close()

    logger.info(
        f"""
        IMPORT FINISHED

        SUCCESS: {success_count}

        ERRORS: {error_count}
        """
    )

    return {
        "success": success_count,
        "errors": error_count,
    }